#!/usr/bin/env python3
"""
Excel Checklist Processor - Handle Excel template uploads and parsing
Supports flexible Excel formats with interactive column mapping
"""

import os
import logging
from typing import Dict, List, Any, Optional, Tuple
import hashlib
from datetime import datetime

# Excel processing imports with fallbacks
try:
    import openpyxl
    from openpyxl import load_workbook
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    logging.warning("openpyxl not available - Excel processing disabled")

logger = logging.getLogger(__name__)

class ExcelProcessor:
    """
    Process Excel checklist templates with flexible column mapping
    """
    
    def __init__(self):
        """Initialize Excel processor"""
        if not EXCEL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel processing")
        
        # Supported file extensions
        self.supported_extensions = ['.xlsx', '.xls', '.xlsm']
        
        # Maximum file size (50MB)
        self.max_file_size = 50 * 1024 * 1024
        
        # Maximum rows to process
        self.max_rows = 10000
    
    def validate_excel_file(self, file_path: str) -> Dict[str, Any]:
        """
        Validate Excel file before processing
        
        Args:
            file_path: Path to Excel file
            
        Returns:
            Dict with validation results
        """
        try:
            # Check file exists
            if not os.path.exists(file_path):
                return {'valid': False, 'error': 'File not found'}
            
            # Check file size
            file_size = os.path.getsize(file_path)
            if file_size > self.max_file_size:
                return {'valid': False, 'error': f'File too large: {file_size / 1024 / 1024:.1f}MB (max: 50MB)'}
            
            # Check file extension
            _, ext = os.path.splitext(file_path.lower())
            if ext not in self.supported_extensions:
                return {'valid': False, 'error': f'Unsupported format: {ext}'}
            
            # Try to load workbook
            try:
                workbook = load_workbook(file_path, read_only=True)
                sheet_names = workbook.sheetnames
                workbook.close()
                
                return {
                    'valid': True,
                    'file_size': file_size,
                    'sheet_names': sheet_names,
                    'sheet_count': len(sheet_names)
                }
                
            except Exception as e:
                return {'valid': False, 'error': f'Invalid Excel file: {str(e)}'}
                
        except Exception as e:
            return {'valid': False, 'error': f'File validation failed: {str(e)}'}
    
    def preview_excel_sheets(self, file_path: str, max_preview_rows: int = 10) -> Dict[str, Any]:
        """
        Preview Excel sheets for column mapping interface
        
        Args:
            file_path: Path to Excel file
            max_preview_rows: Maximum rows to preview per sheet
            
        Returns:
            Dict with sheet previews
        """
        try:
            workbook = load_workbook(file_path, read_only=True)
            sheets_preview = {}
            
            for sheet_name in workbook.sheetnames:
                try:
                    sheet = workbook[sheet_name]
                    
                    # Get sheet dimensions with fallback for problematic files
                    try:
                        sheet_max_row = sheet.max_row if sheet.max_row is not None else 0
                        sheet_max_col = sheet.max_column if sheet.max_column is not None else 0
                        
                        # If sheet reports no dimensions, try to detect actual used range
                        if sheet_max_row == 0 or sheet_max_col == 0:
                            print(f"DEBUG: Sheet '{sheet_name}' reports no dimensions, trying to find actual range...")
                            
                            # Scan for actual content
                            actual_max_row = 0
                            actual_max_col = 0
                            
                            # Check up to 100 rows and 50 columns for content
                            for row_idx in range(1, 101):
                                for col_idx in range(1, 51):
                                    try:
                                        cell = sheet.cell(row=row_idx, column=col_idx)
                                        if cell.value is not None and str(cell.value).strip():
                                            actual_max_row = max(actual_max_row, row_idx)
                                            actual_max_col = max(actual_max_col, col_idx)
                                    except Exception:
                                        continue
                            
                            sheet_max_row = actual_max_row if actual_max_row > 0 else 1
                            sheet_max_col = actual_max_col if actual_max_col > 0 else 1
                            
                            print(f"DEBUG: Found actual dimensions: {sheet_max_row} rows × {sheet_max_col} columns")
                        
                        max_row = min(sheet_max_row, max_preview_rows + 1) if sheet_max_row > 0 else 1
                        max_col = min(sheet_max_col, 20) if sheet_max_col > 0 else 1  # Limit columns for preview
                        
                    except Exception as e:
                        print(f"DEBUG: Error getting sheet dimensions: {e}")
                        max_row = max_preview_rows + 1
                        max_col = 20
                        sheet_max_row = max_row
                        sheet_max_col = max_col
                    
                    # Extract preview data
                    preview_data = []
                    total_non_empty = 0  # Track non-empty cells for debugging
                    
                    for row_idx in range(1, max_row + 1):
                        row_data = []
                        for col_idx in range(1, max_col + 1):
                            try:
                                cell = sheet.cell(row=row_idx, column=col_idx)
                                cell_value = cell.value
                                
                                # Track non-empty cells for debugging
                                if cell_value is not None and str(cell_value).strip():
                                    total_non_empty += 1
                                
                                # Format cell value for display
                                if cell_value is None:
                                    cell_value = ""
                                elif isinstance(cell_value, (int, float)):
                                    cell_value = str(cell_value)
                                elif not isinstance(cell_value, str):
                                    cell_value = str(cell_value)
                                
                                # Safe coordinate access - handle EmptyCell objects
                                try:
                                    coordinate = cell.coordinate
                                    column_letter = cell.column_letter
                                    row_num = cell.row
                                    column_num = cell.column
                                except AttributeError:
                                    # Handle EmptyCell objects or other cell types without these attributes
                                    coordinate = f"{openpyxl.utils.get_column_letter(col_idx)}{row_idx}"
                                    column_letter = openpyxl.utils.get_column_letter(col_idx)
                                    row_num = row_idx
                                    column_num = col_idx
                                
                                row_data.append({
                                    'value': cell_value,
                                    'coordinate': coordinate,
                                    'column_letter': column_letter,
                                    'row': row_num,
                                    'column': column_num
                                })
                                
                            except Exception as e:
                                print(f"DEBUG: Error processing cell at row {row_idx}, col {col_idx}: {e}")
                                # Add a default empty cell if there's an error
                                row_data.append({
                                    'value': "",
                                    'coordinate': f"{openpyxl.utils.get_column_letter(col_idx)}{row_idx}",
                                    'column_letter': openpyxl.utils.get_column_letter(col_idx),
                                    'row': row_idx,
                                    'column': col_idx
                                })
                        
                        preview_data.append(row_data)
                    
                    # Generate column headers (A, B, C, etc.)
                    column_headers = []
                    for col_idx in range(1, max_col + 1):
                        column_headers.append(openpyxl.utils.get_column_letter(col_idx))
                    
                    # Better data detection - check if there are actual non-empty cells
                    has_actual_data = self._sheet_has_actual_data(sheet, preview_data)
                    
                    # If preview data detection failed but we found actual dimensions, check again
                    if not has_actual_data and (sheet_max_row > 1 or sheet_max_col > 1) and total_non_empty > 0:
                        has_actual_data = True
                        print(f"DEBUG: Override - found {total_non_empty} non-empty cells, marking as has_data=True")
                    
                    # Debug logging
                    print(f"DEBUG: Sheet '{sheet_name}' analysis:")
                    print(f"  Original dimensions: {sheet.max_row} × {sheet.max_column}")
                    print(f"  Corrected dimensions: {sheet_max_row} × {sheet_max_col}")
                    print(f"  Preview data rows: {len(preview_data)}")
                    print(f"  Non-empty cells in preview: {total_non_empty}")
                    print(f"  Has actual data detected: {has_actual_data}")
                    
                    sheets_preview[sheet_name] = {
                        'name': sheet_name,
                        'max_row': sheet_max_row,
                        'max_column': sheet_max_col,
                        'preview_rows': len(preview_data),
                        'preview_columns': len(column_headers),
                        'column_headers': column_headers,
                        'data': preview_data,
                        'has_data': has_actual_data,
                        'debug_info': {
                            'sheet_max_row': sheet_max_row,
                            'sheet_max_col': sheet_max_col,
                            'original_max_row': sheet.max_row,
                            'original_max_col': sheet.max_column,
                            'preview_data_rows': len(preview_data),
                            'non_empty_cells_in_preview': total_non_empty,
                            'actual_data_detected': has_actual_data
                        }
                    }
                    
                except Exception as e:
                    logger.error(f"Error previewing sheet '{sheet_name}': {e}")
                    sheets_preview[sheet_name] = {
                        'name': sheet_name,
                        'error': f'Preview failed: {str(e)}',
                        'has_data': False
                    }
            
            workbook.close()
            
            return {
                'success': True,
                'sheets': sheets_preview,
                'total_sheets': len(sheets_preview)
            }
            
        except Exception as e:
            return {
                'success': False,
                'error': f'Excel preview failed: {str(e)}'
            }
    
    def parse_checklist_from_excel(self, 
                                 file_path: str, 
                                 sheets_config: Dict[str, Any]) -> Dict[str, Any]:
        """
        Parse checklist items from Excel file based on configuration
        
        Args:
            file_path: Path to Excel file
            sheets_config: Configuration mapping columns to fields
            
        Returns:
            Dict with parsed checklist items
        """
        try:
            workbook = load_workbook(file_path, read_only=True)
            all_checklist_items = []
            parsing_errors = []
            
            total_items_parsed = 0
            
            # Process each configured sheet
            for sheet_name, config in sheets_config.items():
                if not config.get('enabled', True):
                    continue
                
                try:
                    if sheet_name not in workbook.sheetnames:
                        parsing_errors.append(f"Sheet '{sheet_name}' not found")
                        continue
                    
                    sheet = workbook[sheet_name]
                    sheet_items = self._parse_sheet_items(sheet, sheet_name, config)
                    
                    all_checklist_items.extend(sheet_items['items'])
                    total_items_parsed += sheet_items['count']
                    
                    if sheet_items['errors']:
                        parsing_errors.extend([f"Sheet '{sheet_name}': {err}" for err in sheet_items['errors']])
                
                except Exception as e:
                    error_msg = f"Error parsing sheet '{sheet_name}': {str(e)}"
                    parsing_errors.append(error_msg)
                    logger.error(error_msg)
            
            workbook.close()
            
            # Generate file hash for tracking
            file_hash = self._calculate_file_hash(file_path)
            
            return {
                'success': True,
                'items': all_checklist_items,
                'total_items': total_items_parsed,
                'file_hash': file_hash,
                'sheets_processed': len([s for s in sheets_config.values() if s.get('enabled', True)]),
                'parsing_errors': parsing_errors,
                'parsed_at': datetime.utcnow().isoformat()
            }
            
        except Exception as e:
            return {
                'success': False,
                'error': f'Excel parsing failed: {str(e)}',
                'items': [],
                'total_items': 0
            }
    
    def _parse_sheet_items(self, sheet, sheet_name: str, config: Dict[str, Any]) -> Dict[str, Any]:
        """
        Parse checklist items from a single sheet
        
        Args:
            sheet: openpyxl worksheet object
            sheet_name: Name of the sheet
            config: Configuration for this sheet
            
        Returns:
            Dict with parsed items and metadata
        """
        items = []
        errors = []
        
        try:
            # Get configuration
            column_mapping = config.get('column_mapping', {})
            data_start_row = config.get('data_start_row', 2)
            data_end_row = config.get('data_end_row', sheet.max_row)
            
            # Validate column mappings
            required_fields = ['question_text']
            for field in required_fields:
                if field not in column_mapping or not column_mapping[field]:
                    errors.append(f"Missing required field mapping: {field}")
                    return {'items': items, 'count': 0, 'errors': errors}
            
            # Process data rows
            for row_num in range(data_start_row, min(data_end_row + 1, self.max_rows)):
                try:
                    item_data = self._extract_item_from_row(sheet, row_num, column_mapping, sheet_name)
                    
                    if item_data and item_data.get('question_text', '').strip():
                        items.append(item_data)
                    
                except Exception as e:
                    errors.append(f"Row {row_num}: {str(e)}")
                    continue
            
            return {
                'items': items,
                'count': len(items),
                'errors': errors
            }
            
        except Exception as e:
            errors.append(f"Sheet parsing error: {str(e)}")
            return {'items': items, 'count': 0, 'errors': errors}
    
    def _extract_item_from_row(self, sheet, row_num: int, column_mapping: Dict[str, str], sheet_name: str) -> Dict[str, Any]:
        """
        Extract a single checklist item from a row
        
        Args:
            sheet: openpyxl worksheet object
            row_num: Row number to process
            column_mapping: Column to field mappings
            sheet_name: Name of the sheet
            
        Returns:
            Dict with item data
        """
        item_data = {
            'sheet_name': sheet_name,
            'row_number': row_num,
            'excel_reference': f"{sheet_name}!{row_num}",
        }
        
        # Extract data based on column mappings
        for field_name, column_ref in column_mapping.items():
            if not column_ref:
                continue
            
            try:
                # Convert column reference to cell
                if isinstance(column_ref, str):
                    # Handle column letters (A, B, C) or full references (A1, B2)
                    if column_ref.isalpha():
                        # Just a column letter, combine with row number
                        cell_ref = f"{column_ref}{row_num}"
                    else:
                        # Assume it's a full reference, extract column
                        col_letter = ''.join([c for c in column_ref if c.isalpha()])
                        cell_ref = f"{col_letter}{row_num}"
                else:
                    # Assume it's a column number
                    col_letter = openpyxl.utils.get_column_letter(column_ref)
                    cell_ref = f"{col_letter}{row_num}"
                
                cell_value = sheet[cell_ref].value
                
                # Clean and format the value
                if cell_value is None:
                    cell_value = ""
                elif isinstance(cell_value, (int, float)):
                    cell_value = str(cell_value)
                elif not isinstance(cell_value, str):
                    cell_value = str(cell_value)
                else:
                    cell_value = cell_value.strip()
                
                # Store the value
                item_data[field_name] = cell_value
                
                # Store Excel reference for this field
                item_data[f"{field_name}_excel_ref"] = cell_ref
                
            except Exception as e:
                logger.warning(f"Error extracting {field_name} from {column_ref} at row {row_num}: {e}")
                item_data[field_name] = ""
        
        # Set default values for missing fields
        defaults = {
            'priority': 'medium',
            'mandatory': False,
            'expected_response_type': 'text',
            'requirement_type': 'general',
            'is_active': True,
            'display_order': row_num - 1  # Use row number as order
        }
        
        for field, default_value in defaults.items():
            if field not in item_data or not item_data[field]:
                item_data[field] = default_value
        
        # Convert boolean fields
        boolean_fields = ['mandatory']
        for field in boolean_fields:
            if field in item_data:
                value = str(item_data[field]).lower()
                item_data[field] = value in ['true', 'yes', '1', 'y', 'mandatory', 'required']
        
        return item_data
    
    def _calculate_file_hash(self, file_path: str) -> str:
        """Calculate SHA-256 hash of file for tracking changes"""
        try:
            hash_sha256 = hashlib.sha256()
            with open(file_path, "rb") as f:
                for chunk in iter(lambda: f.read(4096), b""):
                    hash_sha256.update(chunk)
            return hash_sha256.hexdigest()
        except Exception as e:
            logger.error(f"Error calculating file hash: {e}")
            return ""
    
    def get_suggested_column_mappings(self, preview_data: Dict[str, Any]) -> Dict[str, Any]:
        """
        Analyze preview data and suggest column mappings
        
        Args:
            preview_data: Sheet preview data
            
        Returns:
            Dict with suggested mappings
        """
        suggestions = {}
        
        try:
            for sheet_name, sheet_data in preview_data.get('sheets', {}).items():
                if not sheet_data.get('has_data', False):
                    continue
                
                sheet_suggestions = self._analyze_sheet_for_suggestions(sheet_data)
                if sheet_suggestions:
                    suggestions[sheet_name] = sheet_suggestions
            
            return {
                'success': True,
                'suggestions': suggestions
            }
            
        except Exception as e:
            return {
                'success': False,
                'error': f'Column analysis failed: {str(e)}'
            }
    
    def _analyze_sheet_for_suggestions(self, sheet_data: Dict[str, Any]) -> Dict[str, Any]:
        """
        Analyze a single sheet to suggest column mappings
        
        Args:
            sheet_data: Sheet preview data
            
        Returns:
            Dict with suggested column mappings
        """
        suggestions = {}
        
        try:
            data = sheet_data.get('data', [])
            if not data:
                return suggestions
            
            # Analyze header row (first row) for keywords
            header_row = data[0] if data else []
            
            # Keywords to look for in headers
            field_keywords = {
                'question_text': ['question', 'requirement', 'item', 'description', 'text', 'criteria'],
                'section': ['section', 'category', 'area', 'module', 'group'],
                'priority': ['priority', 'importance', 'level', 'criticality'],
                'mandatory': ['mandatory', 'required', 'must', 'compulsory'],
                'category': ['category', 'type', 'class', 'kind'],
                'expected_response_type': ['response', 'answer', 'format', 'type']
            }
            
            # Analyze each column in header row
            for col_idx, cell_data in enumerate(header_row):
                header_text = str(cell_data.get('value', '')).lower().strip()
                column_letter = cell_data.get('column_letter', '')
                
                if not header_text:
                    continue
                
                # Find best matching field for this column
                best_match = None
                best_score = 0
                
                for field_name, keywords in field_keywords.items():
                    score = 0
                    for keyword in keywords:
                        if keyword in header_text:
                            score += len(keyword)  # Longer matches get higher scores
                    
                    if score > best_score:
                        best_score = score
                        best_match = field_name
                
                if best_match and best_score > 0:
                    suggestions[best_match] = {
                        'column': column_letter,
                        'header_text': header_text,
                        'confidence': min(best_score / 10, 1.0)  # Normalize to 0-1
                    }
            
            # Ensure we have at least question_text suggestion
            if 'question_text' not in suggestions and len(header_row) > 0:
                # Use the column with the longest header text as question column
                longest_header_col = max(header_row, 
                                       key=lambda x: len(str(x.get('value', ''))))
                suggestions['question_text'] = {
                    'column': longest_header_col.get('column_letter', 'B'),
                    'header_text': str(longest_header_col.get('value', '')),
                    'confidence': 0.3  # Low confidence fallback
                }
            
            return suggestions
            
        except Exception as e:
            logger.error(f"Error analyzing sheet for suggestions: {e}")
            return {}
    
    def _sheet_has_actual_data(self, sheet, preview_data: List[List[Dict]]) -> bool:
        """
        Check if a sheet has actual non-empty data
        
        Args:
            sheet: Openpyxl worksheet object
            preview_data: Extracted preview data
            
        Returns:
            bool: True if sheet has meaningful data
        """
        try:
            # Method 1: Check if preview data has non-empty cells
            non_empty_cells = 0
            for row in preview_data:
                for cell in row:
                    cell_value = str(cell.get('value', '')).strip()
                    if cell_value and cell_value not in ['', 'None', 'null']:
                        non_empty_cells += 1
            
            # If we have at least 2 non-empty cells, consider it has data
            if non_empty_cells >= 2:
                return True
            
            # Method 2: Check sheet's calculated dimensions
            if sheet.max_row > 1 and sheet.max_column > 0:
                # Try to find any cell with actual content
                for row in sheet.iter_rows(min_row=1, max_row=min(20, sheet.max_row), 
                                         min_col=1, max_col=min(20, sheet.max_column), 
                                         values_only=True):
                    for cell_value in row:
                        if cell_value is not None:
                            cell_str = str(cell_value).strip()
                            if cell_str and cell_str not in ['', 'None', 'null']:
                                return True
            
            # Method 3: Check if there are any formatted cells (might have content)
            try:
                # Check first few rows for any cells with content or formatting
                for row_idx in range(1, min(10, sheet.max_row + 1)):
                    for col_idx in range(1, min(10, sheet.max_column + 1)):
                        cell = sheet.cell(row=row_idx, column=col_idx)
                        if (cell.value is not None or 
                            cell.fill.start_color.index != '00000000' or  # Has background color
                            cell.font.bold or 
                            cell.border.left.style is not None):
                            return True
            except Exception:
                pass  # Ignore formatting check errors
            
            return False
            
        except Exception as e:
            logger.error(f"Error checking if sheet has data: {e}")
            # Default to True to avoid blocking valid sheets
            return True

def create_excel_processor() -> ExcelProcessor:
    """Factory function to create ExcelProcessor instance"""
    try:
        return ExcelProcessor()
    except ImportError:
        raise ImportError("Excel processing requires openpyxl. Install with: pip install openpyxl")

# Export the main class and factory function
__all__ = ['ExcelProcessor', 'create_excel_processor', 'EXCEL_AVAILABLE']