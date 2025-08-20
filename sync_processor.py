#!/usr/bin/env python3
"""
Synchronous Document Processor - replaces Celery tasks
Processes documents immediately without background queue
"""

import os
import logging
from datetime import datetime

def process_document_sync(document):
    """
    Process a document synchronously - replaces Celery background task
    
    Args:
        document: Document model instance
    
    Returns:
        dict: Processing result with success status and any errors
    """
    try:
        print(f"🤖 Processing document: {document.id}")
        
        # Import here to avoid circular imports
        from models import db
        from document_processor import DocumentProcessor
        
        # Update document status to processing
        if hasattr(document, 'processing_status'):
            document.processing_status = 'processing'
            db.session.commit()
        
        # Get filename safely
        filename = getattr(document, 'original_filename', None) or getattr(document, 'filename', 'unknown')
        print(f"📄 Processing file: {filename}")
        
        # Check if file exists
        if not document.file_path or not os.path.exists(document.file_path):
            raise Exception(f"File not found: {document.file_path}")
        
        # Extract content based on file type
        content = ""
        file_ext = filename.lower().split('.')[-1] if '.' in filename else ''
        
        if file_ext == 'txt':
            # Handle text files with multiple encoding attempts
            encodings = ['utf-8', 'latin-1', 'cp1252', 'iso-8859-1']
            for encoding in encodings:
                try:
                    with open(document.file_path, 'r', encoding=encoding, errors='ignore') as f:
                        content = f.read()
                        break
                except (UnicodeDecodeError, Exception):
                    continue
                    
        elif file_ext in ['pdf']:
            # Use document processor for PDF files
            try:
                processor = DocumentProcessor()
                result = processor.extract_text_from_pdf(document.file_path)
                content = result.get('content', '') if isinstance(result, dict) else str(result)
            except Exception as e:
                print(f"⚠️ PDF processing error: {e}")
                content = f"Error processing PDF: {str(e)}"
                
        elif file_ext in ['doc', 'docx']:
            # Use document processor for Word files
            try:
                processor = DocumentProcessor()
                result = processor.extract_text_from_docx(document.file_path)
                content = result.get('content', '') if isinstance(result, dict) else str(result)
            except Exception as e:
                print(f"⚠️ DOCX processing error: {e}")
                content = f"Error processing DOCX: {str(e)}"
                
        elif file_ext in ['xlsx', 'xls']:
            # Handle Excel files
            try:
                import openpyxl
                workbook = openpyxl.load_workbook(document.file_path)
                sheets_content = []
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    sheet_data = []
                    for row in sheet.iter_rows(values_only=True):
                        if any(cell is not None for cell in row):
                            sheet_data.append(' | '.join(str(cell or '') for cell in row))
                    if sheet_data:
                        sheets_content.append(f"Sheet: {sheet_name}\n" + '\n'.join(sheet_data))
                content = '\n\n'.join(sheets_content)
            except Exception as e:
                print(f"⚠️ Excel processing error: {e}")
                content = f"Error processing Excel file: {str(e)}"
        else:
            # Unsupported file type
            content = f"Unsupported file type: {file_ext}"
            print(f"⚠️ Unsupported file type: {file_ext}")
        
        # Update document with extracted content
        if hasattr(document, 'content'):
            document.content = content
        if hasattr(document, 'processing_status'):
            document.processing_status = 'completed'
        if hasattr(document, 'processed_at'):
            document.processed_at = datetime.utcnow()
            
        # Save to database
        db.session.commit()
        
        print(f"✅ Document {document.id} processed successfully")
        print(f"📊 Extracted {len(content)} characters")
        
        # Optional: Trigger any additional analysis
        try:
            from real_analysis_system import RealAnalysisSystem
            analysis_system = RealAnalysisSystem()
            
            # Quick analysis to populate basic fields
            analysis_result = analysis_system.analyze_document_content(
                content=content,
                filename=filename,
                project_id=getattr(document, 'project_id', None)
            )
            
            # Store basic analysis results
            if analysis_result and isinstance(analysis_result, dict):
                if hasattr(document, 'analysis_summary'):
                    document.analysis_summary = analysis_result.get('summary', '')[:500]  # Truncate if too long
                
                db.session.commit()
                print("✅ Basic analysis completed")
                
        except Exception as e:
            print(f"⚠️ Analysis failed (document still processed): {e}")
        
        return {
            'success': True,
            'content_length': len(content),
            'message': f'Document processed successfully - extracted {len(content)} characters'
        }
        
    except Exception as e:
        print(f"❌ Document processing failed: {e}")
        
        # Update document status to failed
        try:
            if hasattr(document, 'processing_status'):
                document.processing_status = 'failed'
                db.session.commit()
        except:
            pass
        
        return {
            'success': False,
            'error': str(e),
            'message': f'Document processing failed: {str(e)}'
        }