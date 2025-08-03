import os
import PyPDF2
import docx
import openpyxl
from PIL import Image
import pytesseract
import anthropic
import logging
from werkzeug.utils import secure_filename

class DocumentProcessor:
    def __init__(self, anthropic_api_key, upload_folder='uploads'):
        self.client = anthropic.Anthropic(api_key=anthropic_api_key)
        self.upload_folder = upload_folder
        self.allowed_extensions = {'pdf', 'doc', 'docx', 'txt', 'xlsx', 'xls', 'ppt', 'pptx'}

    def allowed_file(self, filename):
        return '.' in filename and filename.rsplit('.', 1)[1].lower() in self.allowed_extensions

    def extract_text_from_pdf(self, file_path):
        """Extract text from PDF file"""
        try:
            text = ""
            with open(file_path, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
                for page in pdf_reader.pages:
                    text += page.extract_text() + "\n"
            return text
        except Exception as e:
            logging.error(f"Error extracting text from PDF: {e}")
            return ""

    def extract_text_from_docx(self, file_path):
        """Extract text from DOCX file"""
        try:
            doc = docx.Document(file_path)
            text = ""
            for paragraph in doc.paragraphs:
                text += paragraph.text + "\n"
            return text
        except Exception as e:
            logging.error(f"Error extracting text from DOCX: {e}")
            return ""

    def extract_text_from_xlsx(self, file_path):
        """Extract text from Excel file"""
        try:
            workbook = openpyxl.load_workbook(file_path)
            text = ""
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                text += f"Sheet: {sheet_name}\n"
                for row in sheet.iter_rows(values_only=True):
                    row_text = " | ".join([str(cell) if cell is not None else "" for cell in row])
                    text += row_text + "\n"
            return text
        except Exception as e:
            logging.error(f"Error extracting text from Excel: {e}")
            return ""

    def extract_text_from_file(self, file_path, mime_type):
        """Extract text based on file type"""
        if mime_type == 'application/pdf':
            return self.extract_text_from_pdf(file_path)
        elif mime_type in ['application/vnd.openxmlformats-officedocument.wordprocessingml.document', 'application/msword']:
            return self.extract_text_from_docx(file_path)
        elif mime_type in ['application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', 'application/vnd.ms-excel']:
            return self.extract_text_from_xlsx(file_path)
        elif mime_type == 'text/plain':
            with open(file_path, 'r', encoding='utf-8') as f:
                return f.read()
        else:
            return ""

    def classify_document(self, filename, content):
        """Classify document type using AI"""
        prompt = f"""
        Classify this RFP document based on its filename and content:

        Filename: {filename}
        Content preview: {content[:1000]}...

        Classify into one of these categories:
        - primary: Main RFP document
        - addendum: Updates or changes to the original RFP
        - technical_spec: Technical specifications and requirements
        - commercial_spec: Commercial terms and conditions
        - supporting: Supporting documents, attachments, references
        - qa: Questions and answers, clarifications

        Return only the category name.
        """

        try:
            message = self.client.messages.create(
                model="claude-sonnet-4-20250514",
                max_tokens=50,
                messages=[{"role": "user", "content": prompt}]
            )
            return message.content[0].text.strip().lower()
        except Exception as e:
            logging.error(f"Error classifying document: {e}")
            return "primary"

    def extract_key_points(self, content, document_id):
        """Extract key points from document content"""
        prompt = f"""
        Extract key points from this RFP document and categorize them:

        {content}

        Extract and categorize key points into these types:
        - requirement: Functional or technical requirements
        - constraint: Limitations, restrictions, or conditions
        - deadline: Important dates and deadlines
        - evaluation_criteria: How proposals will be evaluated
        - contact_info: Contact details for queries
        - budget: Cost, pricing, or financial information
        - scope: Project scope and deliverables
        - compliance: Regulatory or compliance requirements
        - deliverable: Expected outputs and deliverables
        - technical_spec: Technical specifications and standards
        - submission: Submission requirements and format

        For each key point, provide:
        1. content: The exact content of the key point
        2. type: One of the categories above
        3. priority: critical, high, medium, or low
        4. page: Page number if identifiable (estimate if needed)
        5. section: Section name if identifiable
        6. confidence: Confidence score between 0.0 and 1.0
        7. tags: Array of relevant tags for easier searching

        Return as JSON array format:
        [
          {{
            "content": "text of the key point",
            "type": "requirement",
            "priority": "high",
            "page": 5,
            "section": "Technical Requirements",
            "confidence": 0.95,
            "tags": ["technical", "mandatory", "system"]
          }}
        ]
        """

        try:
            message = self.client.messages.create(
                model="claude-sonnet-4-20250514",
                max_tokens=4000,
                messages=[{"role": "user", "content": prompt}]
            )

            response_text = message.content[0].text.strip()
            # Clean up response to extract JSON
            if response_text.startswith('```json'):
                response_text = response_text[7:]
            if response_text.endswith('```'):
                response_text = response_text[:-3]

            key_points = json.loads(response_text)
            return key_points
        except Exception as e:
            logging.error(f"Error extracting key points: {e}")
            return []

    def consolidate_key_points(self, all_key_points):
        """Consolidate key points from multiple documents"""
        prompt = f"""
        Analyze and consolidate these key points from multiple RFP documents:

        {json.dumps(all_key_points, indent=2)}

        Tasks:
        1. Group similar/related key points together
        2. Resolve minor conflicts by prioritizing the most recent or authoritative source
        3. Combine complementary information
        4. Create consolidated versions with full traceability

        For each consolidated key point, provide:
        - content: Final consolidated content
        - type: Category type
        - priority: Overall priority level (critical/high/medium/low)
        - source_document_ids: Array of source document IDs
        - source_key_point_ids: Array of original key point IDs that were consolidated
        - final_decision: The consolidated decision/content
        - reasoning: Detailed explanation of why these points were consolidated
        - confidence: Overall confidence score (0.0-1.0)

        Return as JSON array:
        [
          {{
            "content": "consolidated key point text",
            "type": "requirement",
            "priority": "high",
            "source_document_ids": ["doc1", "doc2"],
            "source_key_point_ids": ["kp1", "kp2", "kp3"],
            "final_decision": "Final consolidated decision",
            "reasoning": "These points were consolidated because...",
            "confidence": 0.9
          }}
        ]
        """

        try:
            message = self.client.messages.create(
                model="claude-sonnet-4-20250514",
                max_tokens=4000,
                messages=[{"role": "user", "content": prompt}]
            )

            response_text = message.content[0].text.strip()
            if response_text.startswith('```json'):
                response_text = response_text[7:]
            if response_text.endswith('```'):
                response_text = response_text[:-3]

            consolidated_points = json.loads(response_text)
            return consolidated_points
        except Exception as e:
            logging.error(f"Error consolidating key points: {e}")
            return []

    def detect_conflicts(self, all_key_points):
        """Detect conflicts between key points from different documents"""
        prompt = f"""
        Detect conflicts between these key points from multiple RFP documents:

        {json.dumps(all_key_points, indent=2)}

        Look for conflicts in:
        1. Deadlines (different submission dates, conflicting timelines)
        2. Requirements (contradictory specifications)
        3. Budget information (different amounts, conflicting cost structures)
        4. Scope definitions (conflicting deliverables, different project scope)
        5. Evaluation criteria (different weightings, conflicting evaluation methods)
        6. Contact information (different contacts for the same purpose)
        7. Technical specifications (incompatible technical requirements)
        8. Submission requirements (conflicting format or submission methods)

        For each conflict found:
        - conflict_type: Type of conflict from the list above
        - description: Clear description of what conflicts
        - conflicting_key_point_ids: Array of conflicting key point IDs
        - resolution_strategy: Recommended resolution approach
        - resolution_reasoning: Detailed explanation of why this resolution is recommended

        Return as JSON array:
        [
          {{
            "conflict_type": "deadline",
            "description": "Document A states submission deadline as March 15, while Document B states March 20",
            "conflicting_key_point_ids": ["kp1", "kp5"],
            "resolution_strategy": "latest_document",
            "resolution_reasoning": "Choose the deadline from the most recent document as it likely contains updates"
          }}
        ]
        """

        try:
            message = self.client.messages.create(
                model="claude-sonnet-4-20250514",
                max_tokens=3000,
                messages=[{"role": "user", "content": prompt}]
            )

            response_text = message.content[0].text.strip()
            if response_text.startswith('```json'):
                response_text = response_text[7:]
            if response_text.endswith('```'):
                response_text = response_text[:-3]

            conflicts = json.loads(response_text)
            return conflicts
        except Exception as e:
            logging.error(f"Error detecting conflicts: {e}")
            return []

    def identify_missing_information(self, consolidated_points):
        """Identify missing critical information"""
        prompt = f"""
        Based on these consolidated key points from RFP documents, identify critical missing information:

        {json.dumps(consolidated_points, indent=2)}

        Standard RFP elements to check for:
        1. Submission deadline and format requirements
        2. Contact information for queries and clarifications
        3. Budget constraints or cost expectations
        4. Evaluation criteria and scoring methodology
        5. Project timeline and key milestones
        6. Technical requirements and specifications
        7. Deliverables and acceptance criteria
        8. Compliance and regulatory requirements
        9. Vendor qualification criteria
        10. Contract terms and conditions
        11. Data security and privacy requirements
        12. Support and maintenance requirements
        13. Service level agreements (SLAs)
        14. Integration requirements
        15. Training and documentation requirements

        For each missing item:
        - category: Category of missing information
        - description: What specific information is missing
        - importance: critical, high, medium, or low
        - suggested_questions: Array of specific questions to ask the client

        Return as JSON array:
        [
          {{
            "category": "submission",
            "description": "No clear submission deadline specified",
            "importance": "critical",
            "suggested_questions": [
              "What is the exact deadline for proposal submission?",
              "What timezone should be used for the deadline?",
              "What is the preferred submission format?"
            ]
          }}
        ]
        """

        try:
            message = self.client.messages.create(
                model="claude-sonnet-4-20250514",
                max_tokens=3000,
                messages=[{"role": "user", "content": prompt}]
            )

            response_text = message.content[0].text.strip()
            if response_text.startswith('```json'):
                response_text = response_text[7:]
            if response_text.endswith('```'):
                response_text = response_text[:-3]

            missing_info = json.loads(response_text)
            return missing_info
        except Exception as e:
            logging.error(f"Error identifying missing information: {e}")
            return []
